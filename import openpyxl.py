import openpyxl

def routesum(rl,rr,ct=2,cb=7):
    s=0
    for i in range(rl,rr):
        for j in range(ct,cb):
            cell_obj = sheet_obj.cell(row = i, column = j) 
            s+=cell_obj.value
    return s


path ="Question 1.xlsx"
wb_obj = openpyxl.load_workbook(path) 
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active
i=1
sum=[]
sum.append(routesum(4,9))
sum.append(routesum(14,19))
sum.append(routesum(24,29))
sum.append(routesum(34,39))
m=max(sum)
if(sum.count(m)==1):
    print("Maximum Treasures - ",m,"\nRoute is ",sum.index(m)+1)
else:
        print("Best Treasures -",m)
        j=1
        for i in sum:
            if i==m:
                print("Route no: ",j)
            j+=1

        
"""
while  1:
    cell_obj = sheet_obj.cell(row = j, column = 1) 
    if(cell_obj.value==None):
        print("Number of iterations",i)
        break
    i+=1
print("Maximum Treasures for Route 1")
s=0
sum=[]
for r1 in range(j,i):
    for c1 in range(2,c):
        s+=sheet_obj.cell(row = r1, column = c1).value
        print(sheet_obj.cell(row = r1, column = c1).value,end=" ")
    print()

sum.append(s)    

"""